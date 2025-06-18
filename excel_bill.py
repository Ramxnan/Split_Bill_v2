import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
import os
import time


def number2letter(n):
    return chr(n + 64)

def create_excel(people,items):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font
    wb=Workbook()
    ws=wb.active
    ws.title="Bill Split"

    #Column A should have the names of items

    ws['A1']="Item"
    for i in range(len(items)):
        ws['A'+str(i+2)]=items[i]

    #Column B should have the prices of items
    ws['B1']="Price"

    #Column C should have the Quantity of items
    ws['C1']="Quantity"

    #Column E should have final price of items
    ws['E1']="Final Price"
    for i in range(len(items)):
        ws['E'+str(i+2)]="=B"+str(i+2)+"*C"+str(i+2)

    ws[f'E{len(items)+3}']=f'=SUM(E2:E{len(items)+1})'
    ws[f'E{len(items)+3}'].font=Font(bold=True)
    ws[f'E{len(items)+3}'].alignment=Alignment(horizontal='center')



    #####################################################################################################

    #Row 1 should be bold
    from openpyxl.styles import Font
    ws['A1'].font=Font(bold=True)
    ws['B1'].font=Font(bold=True)
    ws['C1'].font=Font(bold=True)
    ws['D1'].font=Font(bold=True)
    ws['E1'].font=Font(bold=True)

    #Borders for columsn A B C and E
    from openpyxl.styles import Border, Side
    border=Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))

    for i in range(1,len(items)+2):
        ws[f'A{i}'].border=border
        ws[f'A{i}'].alignment=Alignment(horizontal='center')
        ws[f'A{i}'].font=Font(bold=True)
        ws[f'B{i}'].border=border
        ws[f'B{i}'].alignment=Alignment(horizontal='center')
        ws[f'C{i}'].border=border
        ws[f'C{i}'].alignment=Alignment(horizontal='center')
        ws[f'E{i}'].border=border
        ws[f'E{i}'].alignment=Alignment(horizontal='center')
        ws[f'E{i}'].font=Font(bold=True)

    #set row as last row of the excel sheet
    r=ws.max_row
    c=ws.max_column

    r+=2
    c=len(items)+5

    #from row 1 to row r, make the cells black
    from openpyxl.styles import PatternFill
    fill=PatternFill(start_color='000000',end_color='000000',fill_type='solid')

    for i in range(1,c+1):
        ws[f'{number2letter(i)}{r}'].fill=fill

    r=ws.max_row
    c=ws.max_column

    r+=2
    b_r1=r
    #row should have the items from column B
    for i in range(len(items)):
        ws[f'{number2letter(i+2)}{r}']=items[i]
        ws[f'{number2letter(i+2)}{r}'].font=Font(bold=True)
        ws[f'{number2letter(i+2)}{r}'].alignment=Alignment(horizontal='center')

    #Setting cell for PP Total
    c=len(items)+3
    ws[f'{number2letter(c)}{r}']="PP Total"
    ws[f'{number2letter(c)}{r}'].font=Font(bold=True)
    ws[f'{number2letter(c)}{r}'].alignment=Alignment(horizontal='center')

    #column A should have the names of people
    r+=1
    for i in range(len(people)):
        ws[f'A{r+i}']=people[i]
        ws[f'A{r+i}'].font=Font(bold=True)
        ws[f'A{r+i}'].alignment=Alignment(horizontal='center')

        #calculate the sum of each person
        ws[f'{number2letter(c)}{r+i}']=f'=SUM({number2letter(2)}{r+i}:{number2letter(c-2)}{r+i})'
        ws[f'{number2letter(c)}{r+i}'].font=Font(bold=True)
        ws[f'{number2letter(c)}{r+i}'].alignment=Alignment(horizontal='center')


    #Setting cell for Total People sum
    r+=len(people)

    ws[f'{number2letter(c)}{r}']=f'=SUM({number2letter(c)}{b_r1+1}:{number2letter(c)}{r-1})'
    ws[f'{number2letter(c)}{r}'].font=Font(bold=True)
    ws[f'{number2letter(c)}{r}'].alignment=Alignment(horizontal='center')
    fill=PatternFill(start_color='ff4d73',end_color='ff4d73',fill_type='solid')
    ws[f'{number2letter(c)}{r}'].fill=fill


    r+=1
    #column A should have sum
    ws[f'A{r}']="Item Total"
    ws[f'A{r}'].font=Font(bold=True)
    ws[f'A{r}'].alignment=Alignment(horizontal='center')
    for i in range(2,len(items)+2):
        ws[f'{number2letter(i)}{r}']=f'=SUM({number2letter(i)}{b_r1+1}:{number2letter(i)}{r-2})'
        ws[f'{number2letter(i)}{r}'].font=Font(bold=True)
        ws[f'{number2letter(i)}{r}'].alignment=Alignment(horizontal='center')

    #Setting cell for Total Item sum
    ws[f'{number2letter(c-1)}{r}']=f'=SUM(B{r}:{number2letter(c-2)}{r})'
    ws[f'{number2letter(c-1)}{r}'].font=Font(bold=True)
    ws[f'{number2letter(c-1)}{r}'].alignment=Alignment(horizontal='center')
    fill=PatternFill(start_color='ff4d73',end_color='ff4d73',fill_type='solid')
    ws[f'{number2letter(c-1)}{r}'].fill=fill

    r+=1
    #column A should have Balance
    ws[f'A{r}']="Balance"
    ws[f'A{r}'].font=Font(bold=True)
    ws[f'A{r}'].alignment=Alignment(horizontal='center')
    for i in range(2,len(items)+2):
        ws[f'{number2letter(i)}{r}']=f'=E{i}-{number2letter(i)}{r-1}'
        ws[f'{number2letter(i)}{r}'].font=Font(bold=True)
        ws[f'{number2letter(i)}{r}'].alignment=Alignment(horizontal='center')

        
        # Conditional formatting
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

        green_rule = CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, fill=green_fill)
        red_rule = CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill)

        ws.conditional_formatting.add(f'{number2letter(i)}{r}', green_rule)
        ws.conditional_formatting.add(f'{number2letter(i)}{r}', red_rule)




    #####################################################################################################

    #border from b_r1 to r+1 and column A to column number of items+3
    for cell in ws.iter_rows(min_row=b_r1, min_col=1, max_row=r-3, max_col=len(items)+1):
        for c in cell:
            c.border=border

    for cell in ws.iter_rows(min_row=r-1, min_col=1, max_row=r, max_col=len(items)+1):
        for c in cell:
            c.border=border

    for cell in ws.iter_rows(min_row=b_r1, min_col=len(items)+3, max_row=r-2, max_col=len(items)+3):
        for c in cell:
            c.border=border
        

    #Borders for columsn A
    from openpyxl.styles import Border, Side
    border=Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))




    #set row as last row of the excel sheet
    r=ws.max_row
    c=ws.max_column

    r+=2
    c=len(items)+5

    #from row 1 to row r, make the cells black
    from openpyxl.styles import PatternFill
    fill=PatternFill(start_color='000000',end_color='000000',fill_type='solid')


    for i in range(1,c+1):
        ws[f'{number2letter(i)}{r}'].fill=fill

    r=ws.max_row
    c=ws.max_column

    r+=2
    b_r2=r
    #row should have the items from column B
    for i in range(len(items)):
        ws[f'{number2letter(i+2)}{r}']=items[i]
        ws[f'{number2letter(i+2)}{r}'].font=Font(bold=True)
        ws[f'{number2letter(i+2)}{r}'].alignment=Alignment(horizontal='center')


    #column A should have the names of people
    r+=1
    for i in range(len(people)):
        ws[f'A{r+i}']=people[i]
        ws[f'A{r+i}'].font=Font(bold=True)
        ws[f'A{r+i}'].alignment=Alignment(horizontal='center')

    r+=len(people)
    r+=1
    #column A should have sum
    ws[f'A{r}']="Sum"
    ws[f'A{r}'].font=Font(bold=True)
    ws[f'A{r}'].alignment=Alignment(horizontal='center')
    for i in range(2,len(items)+2):
        ws[f'{number2letter(i)}{r}']=f'=SUM({number2letter(i)}{b_r2+1}:{number2letter(i)}{r-2})'
        ws[f'{number2letter(i)}{r}'].font=Font(bold=True)
        ws[f'{number2letter(i)}{r}'].alignment=Alignment(horizontal='center')




    #####################################################################################################

    #border from b_r2 to r+1 and column A to column number of items+3
    for cell in ws.iter_rows(min_row=b_r2, min_col=1, max_row=r-2, max_col=len(items)+1):
        for c in cell:
            c.border=border

    for cell in ws.iter_rows(min_row=r, min_col=1, max_row=r, max_col=len(items)+1):
        for c in cell:
            c.border=border

    #Borders for columsn A
    from openpyxl.styles import Border, Side
    border=Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))


    #set row as last row of the excel sheet
    r=ws.max_row
    c=ws.max_column

    r+=2
    c=len(items)+5

    #from row 1 to row r, make the cells black
    from openpyxl.styles import PatternFill
    fill=PatternFill(start_color='000000',end_color='000000',fill_type='solid')


    for i in range(1,c+1):
        ws[f'{number2letter(i)}{r}'].fill=fill

    #start row for table1 is b_r1
    #start row for table2 is b_r2
    #print(b_r1,b_r2)
    #counter=1
    for i, row in enumerate(range(b_r1 + 1, b_r1 + len(people) + 1)):
        for j, col in enumerate(range(2, len(items) + 2)):

            sum_cell = f'{number2letter(col)}{b_r2 + len(people) + 2}'
            weight_cell = f'{number2letter(col)}{b_r2 + i + 1}'
            #print(i)
            #print(weight_cell)

            weighted_sum_formula = f'{weight_cell}/{sum_cell}*E{j + 2}'

            # Use the IF formula to avoid division by zero
            if_formula = f'=IF({sum_cell}=0, "", {weighted_sum_formula})'

            ws[f'{number2letter(col)}{row}'] = if_formula

            # Optionally color the weight cell
            # fill = PatternFill(start_color='b83853', end_color='b83853', fill_type='solid')
            # ws[weight_cell].fill = fill
        
            

    #set row as last row of the excel sheet
    r=len(items)+5
    c=7


    #from row 1 to row r, make the cells black
    from openpyxl.styles import PatternFill
    fill=PatternFill(start_color='000000',end_color='000000',fill_type='solid')


    for i in range(1,r+1):
        ws[f'{number2letter(c)}{i}'].fill=fill

    c=ws.max_column

    c+=2
    b_r=1
    b_c=c
    r=len(people)+3

    for i in range(len(people)):
        ws[f'{number2letter(c)}{b_r+i+1}']=people[i]

    ws[f'{number2letter(c)}{r}']="Total"

    c+=1
    ws[f'{number2letter(c)}1']="Paid"
    ws[f'{number2letter(c)}{r}']=f'=SUM({number2letter(c)}{b_r+1}:{number2letter(c)}{r-1})'


    c+=1
    ws[f'{number2letter(c)}1']="Split"
    for i in range(2,len(people)+2):
        ws[f'{number2letter(c)}{i}']=f'={number2letter(1+len(items)+2)}{b_r1+i-1}'
    ws[f'{number2letter(c)}{r}']=f'=SUM({number2letter(c)}{b_r+1}:{number2letter(c)}{r-1})'


    c+=1
    ws[f'{number2letter(c)}1']="Owes"
    for i in range(2,len(people)+2):
        ws[f'{number2letter(c)}{i}']=f'={number2letter(c-2)}{i}-{number2letter(c-1)}{i}'
    ws[f'{number2letter(c)}{r}']=f'=SUM({number2letter(c)}{b_r+1}:{number2letter(c)}{r-1})'



    ###############################################################################################

    for cells in ws.iter_rows(min_row=1, min_col=b_c, max_row=len(people)+1, max_col=c):
        for cell in cells:
            cell.border=border
            cell.alignment=Alignment(horizontal='center')
            cell.font=Font(bold=True)

    for cells in ws.iter_rows(min_row=r, min_col=b_c, max_row=r, max_col=c):
        for cell in cells:
            cell.border=border
            cell.alignment=Alignment(horizontal='center')
            cell.font=Font(bold=True)


    return wb

st.title("Bill Splitter")

people_input = st.text_area("Enter names of people (comma-separated)", "adam, bob, charlie, david")
items_input = st.text_area("Enter items (comma-separated)", "item1, item2, item3, item4, item, tax, tips")
file_name = st.text_input("Enter the name of the Excel file", "BillSplit")

#center a button in streamlit

if st.button("Generate Excel"):
    #show a spinner
    with st.spinner("Creating your Excel...    :)"):
        time.sleep(5)
        people = [p.strip() for p in people_input.split(",")]
        items = [i.strip() for i in items_input.split(",")]
        
        workbook = create_excel(people, items)

        #create folder if it doesn't exist
        os.makedirs("Bills", exist_ok=True)

        #save the file in the bills folder
        workbook.save(f"Bills/{file_name}.xlsx")
                
        # with open(file_name + ".xlsx", "rb") as file:
        #     st.download_button(
        #         label="Download Excel",
        #         data=file,
        #         file_name="BillSplit.xlsx",
        #         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        #     )
        #use os to open the file
        os.system(f"start excel Bills/{file_name}.xlsx")

st.write("Please fill in the people's names and items, then press the 'Generate Excel' button to create and download the bill split Excel file.")
